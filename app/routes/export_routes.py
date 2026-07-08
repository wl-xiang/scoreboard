"""结果导出路由：生成精简版 / 详细版 Excel 文件。"""
import io
from urllib.parse import quote
from flask import Blueprint, request, send_file
from ..auth import require_auth
from ..models import Competition
from ..services import compute_leaderboard
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

export_bp = Blueprint('export', __name__)


def _style_sheet(ws, header_count):
    """统一美化表头与单元格样式。"""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2F5597')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center
            cell.border = border
    # 自适应列宽（兼容中文宽度估算）
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ''
            width = sum(2 if ord(ch) > 127 else 1 for ch in val)
            if width > max_len:
                max_len = width
        ws.column_dimensions[letter].width = min(max_len + 4, 50)


@export_bp.route('/competitions/<int:cid>/export', methods=['GET'])
@require_auth
def export_result(cid):
    """导出 Excel 结果：version=simple 精简版 / detailed 详细版。"""
    c = Competition.query.get_or_404(cid)
    version = request.args.get('version', 'simple')
    results, subjects, judge_range = compute_leaderboard(c)

    wb = Workbook()
    ws = wb.active
    ws.title = (c.name or '比赛结果')[:31]

    if version == 'simple':
        # 精简版：名次、选手姓名、最终分数
        ws.append(['名次', '选手姓名', '最终分数'])
        for r in results:
            ws.append([r['rank'], r['player'].name, r['final_score']])
    else:
        # 详细版：名次、编号、姓名、备注 + 各评委分 + 各科目分 + 最终得分
        headers = ['名次', '编号', '姓名', '备注']
        for j in judge_range:
            headers.append('评委%d' % j)
        for s in subjects:
            headers.append(s.name)
        headers.append('最终得分')
        ws.append(headers)
        for r in results:
            p = r['player']
            row = [r['rank'], p.seq, p.name, p.remark]
            for i, _j in enumerate(judge_range):
                row.append(round(r['judge_totals'][i], 4))
            for s in subjects:
                row.append(round(r['subject_totals'].get(s.id, 0), 4))
            row.append(r['final_score'])
            ws.append(row)

    _style_sheet(ws, len(ws[1]))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = '%s_%s结果.xlsx' % (c.name or '比赛', '详细' if version == 'detailed' else '精简')
    return send_file(
        buf, as_attachment=True, download_name=quote(filename),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
